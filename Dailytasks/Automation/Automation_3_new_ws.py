from os.path import dirname, join
import openpyxl                 


excel_doc1 = join(dirname(__file__), 'data', 'TestData1.xlsx')

wb = openpyxl.load_workbook(excel_doc1)

wb.create_sheet('DataSheet2')

print(wb.sheetnames)

print(wb.active)
 
wb.save(excel_doc1)






































