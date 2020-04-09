from os.path import dirname, join
import openpyxl                 
from random import randint

data_folder = join(dirname(__file__), 'data')

for idx in range(1, 6):  
    wb = openpyxl.Workbook()
    doc_name = join(data_folder, 'TestData'+str(idx)+'.xlsx') 
    ws = wb.active
    
    for test_idx in range(1, 5):
        cellA = ws.cell(row=test_idx, column=1)
        cellA.value = 'Test{}:'.format(test_idx)
        
        cellB = ws.cell(row=test_idx, column=2)
        cellB.value = randint(1, 100)
        print(cellB.value)
    
    wb.save(doc_name)
































